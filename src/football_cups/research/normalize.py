from __future__ import annotations

import csv
import hashlib
import io
import json
import zipfile
from collections import Counter
from datetime import UTC, date, datetime, time
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from football_cups.collector.storage import make_run_id

from . import RESEARCH_FLAGS, SCHEMA_VERSION
from .config import ResearchConfig
from .registry import ASSET_BY_ID, ASSETS, ResearchAsset
from .state import ResearchState
from .storage import ResearchStore, stable_id


K1_CSV_SHA256 = "e26210d45df9d691bb81b68c078d494705ddb0aadad73ebc1faae4de36b7a931"
K1_METADATA_SHA256 = "6e7452951c098e30afd47ea2cca729c94b9fe4609011e463ff0e5d3add20d710"
K1_INPUT_HASH = "6285cc00625cb1675881c4c8ec41e8d8938ca5402371d95902809bc3b3344455"


class ResearchNormalizeError(ValueError):
    pass


class ResearchIntegrityError(RuntimeError):
    pass


def _sha256(content: bytes) -> str:
    return hashlib.sha256(content).hexdigest()


def _decimal(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        parsed = Decimal(str(value).strip())
    except (InvalidOperation, ValueError):
        return None
    if not parsed.is_finite():
        return None
    return float(parsed)


def _integer(value: Any) -> int | None:
    parsed = _decimal(value)
    if parsed is None or parsed < 0 or not parsed.is_integer():
        return None
    return int(parsed)


def _decode_csv(content: bytes) -> str:
    for encoding in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ResearchNormalizeError("CSV encoding is unsupported")


def _parse_match_date(value: str) -> date:
    cleaned = value.strip()
    for pattern in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(cleaned, pattern).date()
        except ValueError:
            continue
    raise ResearchNormalizeError(f"unsupported match date: {value!r}")


def _base_record(record_type: str, record_id: str) -> dict[str, Any]:
    return {
        "schema_version": SCHEMA_VERSION,
        "record_type": record_type,
        "record_id": record_id,
        **RESEARCH_FLAGS,
    }


def _source_asset_record(
    asset: ResearchAsset,
    *,
    digest: str,
    size: int,
    blob_path: str,
    observed_at: str | None,
    etag: str | None = None,
    last_modified: str | None = None,
) -> dict[str, Any]:
    observation_key = observed_at or f"legacy-content-{digest}"
    record_id = stable_id("research_source_asset", asset.asset_id, digest, observation_key)
    return {
        **_base_record("ResearchSourceAsset", record_id),
        "source_id": asset.source_id,
        "asset_id": asset.asset_id,
        "url": asset.url,
        "asset_kind": asset.kind,
        "sha256": digest,
        "size_bytes": size,
        "blob_path": blob_path,
        "downloaded_at": observed_at,
        "etag": etag,
        "last_modified": last_modified,
    }


def _fixture_record(
    *,
    source_id: str,
    asset_record_id: str,
    asset_sha256: str,
    fixture_key: str,
    competition: str,
    match_date: date,
    kickoff_time_raw: str | None,
    home: str,
    away: str,
    home_goals: int | None,
    away_goals: int | None,
    result_scope: str,
    result_eligible: bool,
    source_payload: dict[str, Any],
) -> dict[str, Any]:
    record_id = stable_id(
        "research_fixture", source_id, asset_record_id, asset_sha256, fixture_key
    )
    return {
        **_base_record("ResearchFixture", record_id),
        "source_id": source_id,
        "source_asset_record_id": asset_record_id,
        "source_fixture_key": fixture_key,
        "competition": competition,
        "match_date": match_date.isoformat(),
        "kickoff_time_raw": kickoff_time_raw,
        "home_team": home,
        "away_team": away,
        "home_goals": home_goals,
        "away_goals": away_goals,
        "result_scope": result_scope,
        "result_eligible": result_eligible,
        "source_payload": source_payload,
    }


def _market_record(
    fixture: dict[str, Any],
    *,
    asset_sha256: str,
    cohort: str,
    market: str,
    bookmaker: str,
    line: float | None,
    values: dict[str, float],
    contract: str,
) -> dict[str, Any]:
    record_id = stable_id(
        "research_market",
        fixture["record_id"],
        cohort,
        market,
        bookmaker,
        line,
        json.dumps(values, sort_keys=True),
    )
    return {
        **_base_record("ResearchMarketObservation", record_id),
        "fixture_record_id": fixture["record_id"],
        "source_id": fixture["source_id"],
        "asset_sha256": asset_sha256,
        "cohort": cohort,
        "market": market,
        "bookmaker": bookmaker,
        "line": line,
        "values": values,
        "market_contract": contract,
    }


def _quality_event(
    *, source_id: str, fixture_record_id: str, event_type: str, status: str, details: dict[str, Any]
) -> dict[str, Any]:
    record_id = stable_id(
        "research_quality_event", source_id, fixture_record_id, event_type, json.dumps(details, sort_keys=True)
    )
    return {
        **_base_record("ResearchQualityEvent", record_id),
        "source_id": source_id,
        "event_type": event_type,
        "status": status,
        "details": {"fixture_record_id": fixture_record_id, **details},
    }


def _valid_odds(*values: float | None) -> bool:
    return all(value is not None and value > 1.0 for value in values)


def _football_data_markets(
    row: dict[str, str], fixture: dict[str, Any], digest: str, *, contract: str
) -> Iterable[dict[str, Any]]:
    keys = set(row)
    for cohort in ("opening", "closing"):
        if cohort == "opening":
            prefixes = sorted(
                key[:-1]
                for key in keys
                if key.endswith("H")
                and not key.endswith("CH")
                and f"{key[:-1]}D" in keys
                and f"{key[:-1]}A" in keys
                and "AH" not in key
            )
            for prefix in prefixes:
                home, draw, away = (_decimal(row.get(f"{prefix}{suffix}")) for suffix in "HDA")
                if _valid_odds(home, draw, away):
                    yield _market_record(
                        fixture,
                        asset_sha256=digest,
                        cohort=cohort,
                        market="1x2",
                        bookmaker=prefix,
                        line=None,
                        values={"home": home, "draw": draw, "away": away},  # type: ignore[dict-item]
                        contract=contract,
                    )
        else:
            prefixes = sorted(
                key[:-2]
                for key in keys
                if key.endswith("CH")
                and f"{key[:-2]}CD" in keys
                and f"{key[:-2]}CA" in keys
                and "AH" not in key
            )
            for prefix in prefixes:
                home = _decimal(row.get(f"{prefix}CH"))
                draw = _decimal(row.get(f"{prefix}CD"))
                away = _decimal(row.get(f"{prefix}CA"))
                if _valid_odds(home, draw, away):
                    yield _market_record(
                        fixture,
                        asset_sha256=digest,
                        cohort=cohort,
                        market="1x2",
                        bookmaker=prefix,
                        line=None,
                        values={"home": home, "draw": draw, "away": away},  # type: ignore[dict-item]
                        contract=contract,
                    )

    for cohort, marker in (("opening", ">2.5"), ("closing", "C>2.5")):
        for key in sorted(keys):
            if not key.endswith(marker):
                continue
            prefix = key[: -len(marker)]
            if cohort == "opening" and prefix.endswith("C"):
                continue
            under_key = f"{prefix}{'<2.5' if cohort == 'opening' else 'C<2.5'}"
            over, under = _decimal(row.get(key)), _decimal(row.get(under_key))
            if _valid_odds(over, under):
                yield _market_record(
                    fixture,
                    asset_sha256=digest,
                    cohort=cohort,
                    market="total",
                    bookmaker=prefix,
                    line=2.5,
                    values={"over": over, "under": under},  # type: ignore[dict-item]
                    contract=contract,
                )

    for cohort, suffix, line_key in (
        ("opening", "AHH", "AHh"),
        ("closing", "CAHH", "AHCh"),
    ):
        for key in sorted(keys):
            if not key.endswith(suffix):
                continue
            if cohort == "opening" and key.endswith("CAHH"):
                continue
            prefix = key[: -len(suffix)]
            away_key = f"{prefix}{'AHA' if cohort == 'opening' else 'CAHA'}"
            home, away, line = _decimal(row.get(key)), _decimal(row.get(away_key)), _decimal(row.get(line_key))
            if _valid_odds(home, away) and line is not None:
                yield _market_record(
                    fixture,
                    asset_sha256=digest,
                    cohort=cohort,
                    market="asian_handicap",
                    bookmaker=prefix,
                    line=line,
                    values={"home": home, "away": away},  # type: ignore[dict-item]
                    contract=contract,
                )


def normalize_football_data_csv(
    asset: ResearchAsset,
    content: bytes,
    source_asset: dict[str, Any],
    *,
    since: date,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    digest = source_asset["sha256"]
    reader = csv.DictReader(io.StringIO(_decode_csv(content)))
    records: list[dict[str, Any]] = [source_asset]
    counts = Counter()
    for row_number, row in enumerate(reader, start=2):
        home_value = row.get("HomeTeam") or row.get("Home")
        away_value = row.get("AwayTeam") or row.get("Away")
        if not row.get("Date") or not home_value or not away_value:
            counts["missing_identity"] += 1
            continue
        try:
            match_date = _parse_match_date(row["Date"])
        except ResearchNormalizeError:
            counts["invalid_date"] += 1
            continue
        if match_date < since:
            counts["before_since"] += 1
            continue
        competition = (
            asset.competition
            if asset.season and asset.competition
            else row.get("League") or row.get("Div") or asset.competition or "unknown"
        )
        home, away = home_value.strip(), away_value.strip()
        fixture_key = stable_id("source_fixture", competition, match_date, home, away)
        home_goals = _integer(row.get("FTHG") if "FTHG" in row else row.get("HG"))
        away_goals = _integer(row.get("FTAG") if "FTAG" in row else row.get("AG"))
        fixture = _fixture_record(
            source_id=asset.source_id,
            asset_record_id=source_asset["record_id"],
            asset_sha256=digest,
            fixture_key=fixture_key,
            competition=competition,
            match_date=match_date,
            kickoff_time_raw=(row.get("Time") or None),
            home=home,
            away=away,
            home_goals=home_goals,
            away_goals=away_goals,
            result_scope="regular_time_90" if home_goals is not None and away_goals is not None else "unknown",
            result_eligible=home_goals is not None and away_goals is not None,
            source_payload={
                "row_number": row_number,
                "div": row.get("Div"),
                "country": row.get("Country"),
                "league": row.get("League"),
                "season": row.get("Season") or asset.season,
            },
        )
        records.append(fixture)
        contract = "core3_available" if asset.season else "1x2_only"
        markets = list(_football_data_markets(row, fixture, digest, contract=contract))
        records.extend(markets)
        counts["fixtures"] += 1
        counts["market_observations"] += len(markets)
    return records, dict(counts)


def _excel_date(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        return from_excel(value).date()
    return _parse_match_date(str(value))


def _xlsx_size_ok(path: Path, limit: int) -> None:
    if not zipfile.is_zipfile(path):
        raise ResearchIntegrityError("XLSX is not a valid ZIP archive")
    with zipfile.ZipFile(path) as archive:
        if sum(info.file_size for info in archive.infolist()) > limit:
            raise ResearchIntegrityError("XLSX uncompressed size exceeds configured limit")


def normalize_world_cup_xlsx(
    config: ResearchConfig,
    asset: ResearchAsset,
    path: Path,
    source_asset: dict[str, Any],
    *,
    since: date,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    _xlsx_size_ok(path, config.max_xlsx_uncompressed_bytes)
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    records: list[dict[str, Any]] = [source_asset]
    counts = Counter()
    digest = source_asset["sha256"]
    try:
        for sheet_name in ("WorldCup2026", "WorldCup2026Qualifiers"):
            if sheet_name not in workbook.sheetnames:
                raise ResearchNormalizeError(f"required worksheet is missing: {sheet_name}")
            sheet = workbook[sheet_name]
            rows = sheet.iter_rows(values_only=True)
            headers = [str(value).strip() if value is not None else "" for value in next(rows)]
            for row_number, values in enumerate(rows, start=2):
                row = dict(zip(headers, values))
                date_value = row.get("Date")
                home, away = str(row.get("Home") or "").strip(), str(row.get("Away") or "").strip()
                if date_value in (None, "") or not home or not away:
                    continue
                match_date = _excel_date(date_value)
                if match_date < since:
                    continue
                if sheet_name == "WorldCup2026":
                    home_goals, away_goals = _integer(row.get("HGFT")), _integer(row.get("AGFT"))
                    finished = str(row.get("Finished") or "").strip()
                    eligible = finished.lower() == "90 minutes" and home_goals is not None and away_goals is not None
                    result_scope = "regular_time_90" if eligible else "source_full_time_scope_unconfirmed"
                    competition = str(row.get("Competition") or "World Cup 2026")
                else:
                    home_goals, away_goals = _integer(row.get("HG")), _integer(row.get("AG"))
                    finished = ""
                    eligible = False
                    result_scope = "source_full_time_scope_unconfirmed"
                    competition = "World Cup 2026 Qualifiers"
                fixture_key = stable_id("source_fixture", competition, match_date, home, away)
                fixture = _fixture_record(
                    source_id=asset.source_id,
                    asset_record_id=source_asset["record_id"],
                    asset_sha256=digest,
                    fixture_key=fixture_key,
                    competition=competition,
                    match_date=match_date,
                    kickoff_time_raw=str(row.get("Time") or "") or None,
                    home=home,
                    away=away,
                    home_goals=home_goals,
                    away_goals=away_goals,
                    result_scope=result_scope,
                    result_eligible=eligible,
                    source_payload={"sheet": sheet_name, "row_number": row_number, "finished": finished},
                )
                records.append(fixture)
                if not eligible:
                    records.append(
                        _quality_event(
                            source_id=asset.source_id,
                            fixture_record_id=fixture["record_id"],
                            event_type="result_scope_ambiguous",
                            status="failure",
                            details={
                                "sheet": sheet_name,
                                "finished": finished,
                                "reason": "source does not prove a regular-time-only result",
                            },
                        )
                    )
                odds_groups: dict[str, tuple[Any, Any, Any]] = {}
                if sheet_name == "WorldCup2026":
                    for prefix in ("bet365", "Betfair_Exch", "H-Max", "H-Avg"):
                        if prefix.startswith("H-"):
                            label = prefix[2:]
                            odds_groups[label] = (
                                row.get(prefix), row.get(f"D-{label}"), row.get(f"A-{label}")
                            )
                        else:
                            odds_groups[prefix] = (
                                row.get(f"{prefix}-H"), row.get(f"{prefix}-D"), row.get(f"{prefix}-A")
                            )
                else:
                    for label in ("Max", "Avg"):
                        odds_groups[label] = (
                            row.get(f"H_{label}"), row.get(f"D_{label}"), row.get(f"A_{label}")
                        )
                for bookmaker, values in odds_groups.items():
                    home_odd, draw_odd, away_odd = (_decimal(value) for value in values)
                    if _valid_odds(home_odd, draw_odd, away_odd):
                        records.append(
                            _market_record(
                                fixture,
                                asset_sha256=digest,
                                cohort="closing",
                                market="1x2",
                                bookmaker=bookmaker,
                                line=None,
                                values={"home": home_odd, "draw": draw_odd, "away": away_odd},  # type: ignore[dict-item]
                                contract="1x2_only",
                            )
                        )
                        counts["market_observations"] += 1
                counts["fixtures"] += 1
                if not eligible:
                    counts["result_scope_ambiguous"] += 1
    finally:
        workbook.close()
    return records, dict(counts)


def _asset_from_legacy_name(name: str) -> ResearchAsset | None:
    stem = Path(name).stem
    parts = stem.split("-")
    if len(parts) != 2:
        return None
    season_code, league_code = parts
    return next(
        (
            asset
            for asset in ASSETS
            if asset.url.endswith(f"/{season_code}/{league_code}.csv")
        ),
        None,
    )


def normalize_available_assets(
    config: ResearchConfig, *, since: date
) -> dict[str, Any]:
    store = ResearchStore(config)
    state = ResearchState(config.state_path)
    run_id = make_run_id()
    candidates: dict[str, tuple[ResearchAsset, Path, dict[str, Any]]] = {}
    try:
        for cache in state.all_assets():
            asset = ASSET_BY_ID.get(str(cache["asset_id"]))
            if not asset or not cache.get("blob_path"):
                continue
            path = config.research_dir / str(cache["blob_path"])
            if path.is_file():
                candidates[asset.asset_id] = (asset, path, cache)

        legacy_root = config.research_dir / "raw" / "football-data"
        for path in sorted(legacy_root.rglob("*.csv")) if legacy_root.is_dir() else []:
            asset = _asset_from_legacy_name(path.name)
            if not asset or asset.asset_id in candidates:
                continue
            content = path.read_bytes()
            digest, blob_path = store.store_blob(content, "csv")
            candidates[asset.asset_id] = (
                asset,
                blob_path,
                {
                    "sha256": digest,
                    "blob_path": blob_path.relative_to(config.research_dir).as_posix(),
                    "observed_at": None,
                    "etag": None,
                    "last_modified": None,
                },
            )

        summaries: dict[str, Any] = {}
        for asset_id, (asset, path, cache) in sorted(candidates.items()):
            content = path.read_bytes()
            digest = _sha256(content)
            if digest != cache["sha256"]:
                raise ResearchIntegrityError(f"cached asset hash changed: {asset_id}")
            source_asset = _source_asset_record(
                asset,
                digest=digest,
                size=len(content),
                blob_path=path.relative_to(config.research_dir).as_posix(),
                observed_at=cache.get("observed_at"),
                etag=cache.get("etag"),
                last_modified=cache.get("last_modified"),
            )
            if asset.kind == "football_data_csv":
                records, summary = normalize_football_data_csv(
                    asset, content, source_asset, since=since
                )
            elif asset.kind == "world_cup_xlsx":
                records, summary = normalize_world_cup_xlsx(
                    config, asset, path, source_asset, since=since
                )
            else:
                continue
            store.write_records(asset.source_id, run_id, asset.asset_id, records)
            summaries[asset_id] = summary
        store.write_manifest(
            run_id,
            "normalize",
            {
                "schema_version": 1,
                "run_id": run_id,
                "status": "completed",
                "since": since.isoformat(),
                "assets": summaries,
            },
        )
        return {"run_id": run_id, "assets": summaries}
    finally:
        state.close()


def import_k1_dataset(
    config: ResearchConfig, csv_path: Path, metadata_path: Path
) -> dict[str, Any]:
    csv_content = csv_path.read_bytes()
    metadata_content = metadata_path.read_bytes()
    if _sha256(csv_content) != K1_CSV_SHA256:
        raise ResearchIntegrityError("K1 CSV SHA-256 does not match the accepted artifact")
    if _sha256(metadata_content) != K1_METADATA_SHA256:
        raise ResearchIntegrityError("K1 metadata SHA-256 does not match the accepted artifact")
    metadata = json.loads(metadata_content.decode("utf-8"))
    if metadata.get("inputHash") != K1_INPUT_HASH:
        raise ResearchIntegrityError("K1 metadata inputHash does not match")
    rows = list(csv.DictReader(io.StringIO(csv_content.decode("utf-8-sig"))))
    fixture_ids = [str(row.get("fixture_id") or "") for row in rows]
    seasons = Counter(str(row.get("season") or "") for row in rows)
    if len(rows) != 330 or len(set(fixture_ids)) != 330 or seasons != Counter({"2025": 228, "2026": 102}):
        raise ResearchNormalizeError("K1 row, fixture, or season counts do not match the contract")

    store = ResearchStore(config)
    csv_digest, csv_blob = store.store_blob(csv_content, "csv")
    metadata_digest, metadata_blob = store.store_blob(metadata_content, "json")
    run_id = make_run_id()
    source_asset_id = stable_id("research_source_asset", "k1-derived-core3", csv_digest)
    records: list[dict[str, Any]] = [
        {
            **_base_record("ResearchSourceAsset", source_asset_id),
            "source_id": "k1-derived-core3",
            "asset_id": "k1-core3-features-2025-2026",
            "url": None,
            "asset_kind": "derived_feature_dataset",
            "sha256": csv_digest,
            "size_bytes": len(csv_content),
            "blob_path": csv_blob.relative_to(config.research_dir).as_posix(),
            "metadata_sha256": metadata_digest,
            "metadata_blob_path": metadata_blob.relative_to(config.research_dir).as_posix(),
            "input_hash": K1_INPUT_HASH,
            "downloaded_at": None,
        }
    ]
    for row in rows:
        fixture_id = row["fixture_id"]
        record_id = stable_id("research_feature_row", csv_digest, fixture_id)
        records.append(
            {
                **_base_record("ResearchFeatureRow", record_id),
                "source_id": "k1-derived-core3",
                "source_asset_record_id": source_asset_id,
                "source_fixture_key": fixture_id,
                "competition": "K1",
                "match_date": str(row["kickoff"])[:10],
                "season": row["season"],
                "cohort": "derived_closing_features",
                "feature_schema": row.get("feature_schema"),
                "market_contract": row.get("market_contract"),
                "input_hash": row.get("input_hash"),
                "result_scope": "regular_time_90",
                "result_eligible": True,
                "features": row,
            }
        )
    store.write_records("k1-derived-core3", run_id, "k1-core3-features", records)
    store.write_manifest(
        run_id,
        "k1-import",
        {
            "schema_version": 1,
            "run_id": run_id,
            "status": "completed",
            "csv_sha256": csv_digest,
            "metadata_sha256": metadata_digest,
            "input_hash": K1_INPUT_HASH,
            "rows": 330,
            "seasons": dict(sorted(seasons.items())),
        },
    )
    return {"run_id": run_id, "rows": 330, "csv_sha256": csv_digest}
